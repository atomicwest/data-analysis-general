import openpyxl
from openpyxl import load_workbook
from openpyxl.comments import Comment

filename = "donut-info-colors.xlsx"

#read_only needs to be False in order to read comments
wb = load_workbook(filename=filename, read_only=False)
ws = wb['One-hundred Donuts']

#print individual properties
print(ws['A2'].value)
print(ws['A2'].comment)

#print values and comments
for l in ['A', 'B', 'C', 'D']:
    for n in range(1,100):
        cell = l+str(n)
        
        v = ws[cell].value
        try:
            co = ws[cell].comment
        except:
            co = "no comment"
        
        print("%s || %s" % (v, co))

'''
for row in ws.rows:
    
    # print type(row)
    # for cell in row:
    #     print(cell.value)
    # row = ""
    for cell in row:
        #print(type(cell.value))
        print(cell.value)
        print(cell.comment)
        # try:
        #     print("comment: %s" % cell.comment)
        # except:
        #     print("no comment")
        # print(cell.comment)
        # row+=str(cell.value)
    # print(row)

'''